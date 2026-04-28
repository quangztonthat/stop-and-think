"""Apply GLOSSARY.xlsx -> 18 book HTML files.
v2 schema: Term | Display | Vietnamese | Source | Vol | Category | Confidence | Notes

Usage: python3 _apply_glossary.py [--dry-run] [--include-wildcard]
"""
import os
import re
import sys
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.abspath(__file__))
GLOSSARY_PATH = os.path.join(ROOT, 'GLOSSARY.xlsx')
BOOKS_DIR = os.path.join(ROOT, 'books')

DRY_RUN = '--dry-run' in sys.argv
INCLUDE_WILDCARD = '--include-wildcard' in sys.argv


def load_glossary():
    wb = load_workbook(GLOSSARY_PATH, data_only=True)
    ws = wb['Glossary']
    rows = []
    skipped_wildcard = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        term = row[0]
        display = row[1]
        source = row[3] if len(row) > 3 else ''
        if not term or not display:
            continue
        if str(term).strip() == str(display).strip():
            continue
        src = str(source).strip() if source else ''
        if src == '*' and not INCLUDE_WILDCARD:
            skipped_wildcard += 1
            continue
        rows.append((str(term).strip(), str(display).strip(), src))
    print(f'  Skipped {skipped_wildcard} wildcard "*" terms (use --include-wildcard to apply them)', flush=True)
    return rows


def apply_to_html(filepath, glossary_rows):
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    original = content
    changes = {}
    sorted_rows = sorted(glossary_rows, key=lambda r: -len(r[0]))
    for term, display, source in sorted_rows:
        escaped = re.escape(term)
        if ' ' in term or "'" in term or '-' in term:
            pattern = escaped
        else:
            pattern = r'\b' + escaped + r'\b'
        matches = re.findall(pattern, content)
        if not matches:
            continue
        new_content = re.sub(pattern, display.replace('\\', r'\\'), content)
        if new_content != content:
            content = new_content
            changes[term] = len(matches)
    return changes, content if content != original else None


def main():
    if not os.path.exists(GLOSSARY_PATH):
        print(f'ERROR: {GLOSSARY_PATH} not found', flush=True)
        sys.exit(1)
    print(f'Reading {GLOSSARY_PATH}...', flush=True)
    glossary = load_glossary()
    print(f'  Loaded {len(glossary)} term(s) with non-trivial display', flush=True)
    if not glossary:
        return

    files = sorted([f for f in os.listdir(BOOKS_DIR) if f.endswith('.html')])
    total_files_changed = 0
    total_replacements = 0

    for f in files:
        path = os.path.join(BOOKS_DIR, f)
        changes, new_content = apply_to_html(path, glossary)
        if changes:
            n = sum(changes.values())
            total_files_changed += 1
            total_replacements += n
            mode = 'DRY' if DRY_RUN else 'WRITE'
            print(f'  [{mode}] {f}: {n} replacement(s) across {len(changes)} term(s)', flush=True)
            for term, count in sorted(changes.items(), key=lambda x: -x[1])[:5]:
                print(f'         {count}x "{term}"', flush=True)
            if not DRY_RUN and new_content:
                with open(path, 'w', encoding='utf-8') as fp:
                    fp.write(new_content)

    print('', flush=True)
    print(f'=== SUMMARY ===', flush=True)
    print(f'  Files changed:  {total_files_changed} / {len(files)}', flush=True)
    print(f'  Replacements:   {total_replacements}', flush=True)
    print(f'  Mode:           {"DRY RUN" if DRY_RUN else "APPLIED"}', flush=True)


if __name__ == '__main__':
    main()
